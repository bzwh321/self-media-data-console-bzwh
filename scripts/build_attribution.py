# -*- coding: utf-8 -*-
"""构建粉丝增长归因分析数据。

数据源：
- 小红书：self_media_content_detail.csv（platform=xhs，含 new_followers 涨粉字段）
- 抖音：monthly/content_list/*.xlsx（含"粉丝增量"字段，按作品取最新快照）
- B站/公众号：平台后台不提供单内容涨粉归因，注明排除

输出：runtime-data/console-state/attribution.json

归因方法（基于真实采集字段，非 AI 生成结论）：
1. top_contents: 按单内容 new_followers 降序 Top N，计算涨粉贡献占比
2. by_type: 按体裁聚合涨粉，看哪类内容涨粉效率高
3. by_platform: 按平台聚合涨粉
"""
from __future__ import annotations

import csv
import json
import os
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl 未安装，请运行 py -m pip install openpyxl", file=sys.stderr)
    sys.exit(1)

DATA_ROOT = Path(os.environ.get("SELF_MEDIA_DATA_ROOT", str(PROJECT_ROOT / "sample-data" / "self-media")))
DASHBOARD_DIR = DATA_ROOT / "dashboard-normalized"
DETAIL_CSV = DASHBOARD_DIR / "self_media_content_detail.csv"
DOUYIN_MONTHLY_DIR = DATA_ROOT / "抖音数据" / "服务器同步" / "抖音+高清发布" / "monthly" / "content_list"
DOUYIN_RAW_DIR = DATA_ROOT / "抖音数据" / "服务器同步" / "抖音+高清发布" / "raw" / "content_list"

PROJECT_ROOT = Path(__file__).resolve().parents[1]
OUTPUT_PATH = PROJECT_ROOT / "runtime-data" / "console-state" / "attribution.json"

TOP_N = 20


def number(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        s = str(v).strip().replace(",", "").replace("¥", "").replace("%", "")
        if s in {"", "-", "--", "nan", "None"}:
            return 0.0
        return float(s)
    except (TypeError, ValueError):
        return 0.0


def parse_dt(v) -> str:
    """解析时间为 ISO 字符串，用于比较快照新旧。"""
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return v.isoformat(timespec="seconds")
    s = str(v).strip()
    try:
        return datetime.fromisoformat(s).isoformat(timespec="seconds")
    except ValueError:
        pass
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).isoformat(timespec="seconds")
        except ValueError:
            continue
    return s


def iso_day(v) -> str:
    iso = parse_dt(v)
    return iso[:10] if iso else ""


def clean(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in {"nan", "none"} else s


def load_xhs_attribution() -> dict:
    """从小红书 content_detail.csv 读取，按作品取最新快照的 new_followers。"""
    if not DETAIL_CSV.exists():
        return {"contents": [], "total_new_followers": 0}
    latest_by_key = {}  # (account,title,publish) -> (snapshot_iso, row)
    with open(DETAIL_CSV, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row.get("platform") != "xhs":
                continue
            title = clean(row.get("content_title"))
            publish = clean(row.get("publish_time"))
            if not title or not publish:
                continue
            key = (title, publish)  # 不区分账号（小红书两个账号内容镜像，按内容去重）
            snap_iso = parse_dt(row.get("snapshot_time") or row.get("source_mtime"))
            prev = latest_by_key.get(key)
            if prev is None or snap_iso > prev[0]:
                latest_by_key[key] = (snap_iso, row)
    contents = []
    total_nf = 0.0
    for snap_iso, row in latest_by_key.values():
        nf = number(row.get("new_followers"))
        if nf <= 0:
            continue
        title = clean(row.get("content_title"))
        views = number(row.get("views"))
        likes = number(row.get("likes"))
        comments = number(row.get("comments"))
        favorites = number(row.get("favorites"))
        total_nf += nf
        contents.append({
            "platform": "xhs",
            "platform_name": "小红书",
            "title": title,
            "publish_date": iso_day(row.get("publish_time")),
            "content_type": clean(row.get("content_type")) or "未知",
            "new_followers": nf,
            "views": views,
            "likes": likes,
            "comments": comments,
            "favorites": favorites,
            "interactions": likes + comments + favorites,
            "interact_rate": round((likes + comments + favorites) / views, 4) if views > 0 else 0.0,
            "content_url": clean(row.get("content_url")),
            "content_id": clean(row.get("content_id")),
            "source_file": clean(row.get("source_file")),
        })
    return {"contents": contents, "total_new_followers": total_nf}


def load_douyin_attribution() -> dict:
    """从抖音 content_list 读取，按作品取最新 source_mtime 的粉丝增量。

    合并 monthly 聚合文件 + raw 分片（raw 含最新月份），按 (title, publish)
    去重取最新快照，确保归因覆盖到最新月份数据。
    """
    monthly_files = sorted(DOUYIN_MONTHLY_DIR.glob("*.xlsx")) if DOUYIN_MONTHLY_DIR.exists() else []
    raw_files = sorted(DOUYIN_RAW_DIR.rglob("*.xlsx")) if DOUYIN_RAW_DIR.exists() else []
    xlsx_files = monthly_files + raw_files
    if not xlsx_files:
        return {"contents": [], "total_new_followers": 0}
    latest_by_key = {}  # (title,publish) -> (source_mtime_iso, row_dict)
    for xlsx in xlsx_files:
        try:
            wb = load_workbook(xlsx, read_only=True, data_only=True)
        except Exception as e:
            print(f"WARN: 读取 {xlsx.name} 失败: {e}", file=sys.stderr)
            continue
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            continue
        headers = [str(c).strip() if c is not None else "" for c in rows[0]]
        # 定位列
        col = {h: i for i, h in enumerate(headers)}
        idx_title = col.get("作品名称", -1)
        idx_publish = col.get("发布时间", -1)
        idx_nf = col.get("粉丝增量", -1)
        idx_views = col.get("播放量", -1)
        idx_likes = col.get("点赞量", -1)
        idx_comments = col.get("评论量", -1)
        idx_favorites = col.get("收藏量", -1)
        idx_shares = col.get("分享量", -1)
        idx_type = col.get("体裁", -1)
        idx_snap = col.get("快照采集时间", -1)
        idx_src_mtime = col.get("source_mtime", -1)
        if idx_title < 0 or idx_publish < 0 or idx_nf < 0:
            continue
        for r in rows[1:]:
            if not r or not any(c is not None and str(c).strip() for c in r):
                continue
            title = clean(r[idx_title])
            publish = clean(r[idx_publish])
            if not title or not publish:
                continue
            snap_val = ""
            if idx_snap >= 0:
                snap_val = clean(r[idx_snap])
            if not snap_val and idx_src_mtime >= 0:
                snap_val = clean(r[idx_src_mtime])
            snap_iso = parse_dt(snap_val)
            key = (title, publish)
            prev = latest_by_key.get(key)
            if prev is None or snap_iso > prev[0]:
                latest_by_key[key] = (snap_iso, r, xlsx.name, col)
    contents = []
    total_nf = 0.0
    for snap_iso, r, src_name, col in latest_by_key.values():
        i_nf = col.get("粉丝增量", -1)
        i_views = col.get("播放量", -1)
        i_likes = col.get("点赞量", -1)
        i_comments = col.get("评论量", -1)
        i_favorites = col.get("收藏量", -1)
        i_type = col.get("体裁", -1)
        i_title = col.get("作品名称", -1)
        i_publish = col.get("发布时间", -1)
        i_url = col.get("作品链接", -1)
        i_id = col.get("内容唯一键", col.get("作品ID", -1))
        nf = number(r[i_nf]) if i_nf >= 0 else 0
        if nf <= 0:
            continue
        views = number(r[i_views]) if i_views >= 0 else 0
        likes = number(r[i_likes]) if i_likes >= 0 else 0
        comments = number(r[i_comments]) if i_comments >= 0 else 0
        favorites = number(r[i_favorites]) if i_favorites >= 0 else 0
        total_nf += nf
        contents.append({
            "platform": "douyin",
            "platform_name": "抖音",
            "title": clean(r[i_title]) if i_title >= 0 else "",
            "publish_date": iso_day(r[i_publish]) if i_publish >= 0 else "",
            "content_type": clean(r[i_type]) if i_type >= 0 else "未知",
            "new_followers": nf,
            "views": views,
            "likes": likes,
            "comments": comments,
            "favorites": favorites,
            "interactions": likes + comments + favorites,
            "interact_rate": round((likes + comments + favorites) / views, 4) if views > 0 else 0.0,
            "content_url": clean(r[i_url]) if i_url >= 0 else "",
            "content_id": clean(r[i_id]) if i_id >= 0 else "",
            "source_file": src_name,
        })
    return {"contents": contents, "total_new_followers": total_nf}


def build_attribution() -> dict:
    xhs = load_xhs_attribution()
    dy = load_douyin_attribution()

    all_contents = xhs["contents"] + dy["contents"]
    # Top N 按涨粉降序
    all_contents.sort(key=lambda c: c["new_followers"], reverse=True)
    top = all_contents[:TOP_N]
    grand_total = xhs["total_new_followers"] + dy["total_new_followers"]
    for c in top:
        c["contribution_pct"] = round(c["new_followers"] / grand_total * 100, 2) if grand_total > 0 else 0.0

    # 按体裁聚合
    by_type_map = defaultdict(lambda: {"total_new_followers": 0.0, "count": 0, "total_views": 0.0})
    for c in all_contents:
        t = c["content_type"] or "未知"
        by_type_map[t]["total_new_followers"] += c["new_followers"]
        by_type_map[t]["count"] += 1
        by_type_map[t]["total_views"] += c["views"]
    by_type = []
    for t, v in sorted(by_type_map.items(), key=lambda kv: kv[1]["total_new_followers"], reverse=True):
        by_type.append({
            "content_type": t,
            "total_new_followers": round(v["total_new_followers"]),
            "count": v["count"],
            "avg_new_followers": round(v["total_new_followers"] / v["count"], 1) if v["count"] else 0,
            "total_views": round(v["total_views"]),
        })

    # 按平台聚合
    by_platform = [
        {
            "platform": "xhs",
            "platform_name": "小红书",
            "total_new_followers": round(xhs["total_new_followers"]),
            "content_count": len(xhs["contents"]),
            "excluded": False,
        },
        {
            "platform": "douyin",
            "platform_name": "抖音",
            "total_new_followers": round(dy["total_new_followers"]),
            "content_count": len(dy["contents"]),
            "excluded": False,
        },
        {
            "platform": "bili",
            "platform_name": "B站",
            "excluded": True,
            "reason": "B站后台不提供单内容涨粉归因，仅账号级日涨粉趋势",
        },
        {
            "platform": "wechat",
            "platform_name": "公众号",
            "excluded": True,
            "reason": "公众号后台不提供文章级涨粉归因",
        },
    ]

    return {
        "schema": "attribution.v1",
        "generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
        "method": "基于内容级 new_followers 字段（小红书涨粉/抖音粉丝增量），按最新快照取值，非AI生成",
        "grand_total_new_followers": round(grand_total),
        "top_contents": top,
        "by_type": by_type,
        "by_platform": by_platform,
        "excluded_platforms": ["bili", "wechat"],
    }


def main():
    result = build_attribution()
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"归因数据已生成: {OUTPUT_PATH}")
    print(f"  总涨粉(小红书+抖音): {result['grand_total_new_followers']}")
    print(f"  Top内容数: {len(result['top_contents'])}")
    print(f"  体裁分组数: {len(result['by_type'])}")
    for p in result["by_platform"]:
        if p.get("excluded"):
            print(f"  {p['platform_name']}: 排除（{p['reason']}）")
        else:
            print(f"  {p['platform_name']}: 涨粉{p['total_new_followers']} / 内容{p['content_count']}条")


if __name__ == "__main__":
    main()
